"""Shared workbook sheets for motility, crawling and colony survival.

Adds the layer those three were missing — README, run_info, plate_summary,
condition_summary, qc — in the same order and with the same wording style as
development_results.xlsx, so a reader who has seen one workbook can read the
others without relearning anything. Every sheet a pipeline already wrote is left
alone; see the "Why the new names" row of readme_common for why these are not
called per_plate / per_condition.

VALUES, NOT FORMULAS. survival_excel writes live formulas and then verifies them
against an expected-value plan after recalculation. That machinery is worth its
cost for the assay whose numbers are the paper's headline; replicating it for
three more pipelines would triple the surface area for a class of bug (a wrong
cell reference) that only exists because the formulas exist. These sheets carry
computed values, and the CSV beside them is the same numbers, so a disagreement
between two views of the same run is not possible here.

WHAT n MEANS, EVERY TIME. condition_summary rows are the mean POOLED OVER ITEMS
with ±1 SEM, and <metric>_n is the n. The plates of a condition are not
replicates of each other, so their means are not averaged; the old across-plate
SD survives as <metric>_plate_spread_sd, a QC column and never an error bar. The
pooled median is carried alongside in its own column rather than mixed in,
because the two answer different questions and a reader who does not know which
one they are looking at is worse off than one who sees both.
"""
from __future__ import annotations

from datetime import datetime
from typing import Callable, Optional, Sequence

import assay_common as AC

_FONT = "Calibri"


def _style_header(ws, n_cols: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor="EFEFEC")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=_FONT, size=10, bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, headers: Sequence[str], overrides: Optional[dict] = None) -> None:
    from openpyxl.utils import get_column_letter
    overrides = overrides or {}
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = overrides.get(
            h, min(34, max(10, len(str(h)) + 3)))


def _sheet(wb, name: str, headers: Sequence[str], rows: Sequence[Sequence],
           overrides: Optional[dict] = None):
    from openpyxl.styles import Font
    ws = wb.create_sheet(name)
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    _style_header(ws, len(headers))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = Font(name=_FONT, size=10)
    _autosize(ws, headers, overrides)
    return ws


def write_readme(wb, assay: str, lines: Sequence[tuple]) -> None:
    from openpyxl.styles import Alignment, Font
    ws = wb.create_sheet("README", 0)
    ws.append([f"WHAT THIS WORKBOOK IS — {assay}", ""])
    ws.append(["", ""])
    for k, v in lines:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 118
    ws.cell(row=1, column=1).font = Font(name=_FONT, size=11, bold=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row:
            if not cell.font.bold:
                cell.font = Font(name=_FONT, size=10)


def write_run_info(wb, rows: Sequence[tuple]) -> None:
    _sheet(wb, "run_info", ["key", "value"], list(rows),
           overrides={"key": 34, "value": 110})


def readme_common(metrics: Sequence[AC.Metric], keep_note: str,
                  unit_note: str, timepoints: Sequence = ()) -> list:
    """The paragraphs every one of these workbooks needs, in one place."""
    rows = []
    if len(timepoints or ()) > 1:
        rows.append((
            "This run is a TIMECOURSE",
            "Every plate and condition sheet carries a timepoint_h column and "
            "one row PER TIMEPOINT: " + ", ".join(f"{t:g} h" for t in timepoints)
            + ". A row is identified by (timepoint_h, condition) — the "
              "condition name alone appears once per imaging day, and reading a "
              "number without its timepoint means reading an arbitrary day. "
              "Nothing is averaged across days anywhere in this workbook, "
              "including % of control on the qc sheet, which uses each day's "
              "own control."))
    rows += [
        ("Sheets",
         "plate_summary is one row per plate; condition_summary is one row per "
         "condition; qc is what there was to measure and how much survived; "
         "dist_histogram and dist_summary are the distribution the figure "
         "draws. Every sheet this workbook already carried is unchanged, "
         "including any earlier per_condition — see the next row."),
        ("Why the new names",
         "plate_summary and condition_summary rather than per_plate and "
         "per_condition, because two of these pipelines already ship sheets "
         "with those names and one of them aggregates worms instead of plates. "
         "Renaming a sheet under someone's script would be worse than an extra "
         "sheet. Where both exist, these are the ones the figures and the "
         "explorer use."),
        ("The replication unit is the ITEM, not the plate",
         "condition_summary pools every item of a condition at a timepoint and "
         "takes the mean over them: <metric>_mean, with <metric>_n as the n. "
         "The plates of one condition are not replicates of each other — they "
         "exist so more animals could be imaged in one experiment — so their "
         "means are not averaged. A plate with 200 animals now carries the "
         "weight of 200, and adding a plate adds animals rather than a vote. "
         "plate_summary still reports each plate on its own."),
        ("SEM is the error bar, SD is the spread",
         "<metric>_sd is the sample SD (ddof=1) over the items, i.e. how "
         "different the animals were from each other. <metric>_sem is that SD "
         "over the root of <metric>_n and is what the figures and the explorer "
         "draw. SEM says how well this condition's mean is pinned by the "
         "animals imaged; it is NOT replicate error and must not be read as "
         "how far the number would move if the experiment were repeated — one "
         "experiment cannot tell you that. A condition with a single item has "
         "no spread, so its SD and SEM are blank rather than 0.0."),
        ("…and the QC columns",
         "<metric>_pooled_median is the median over the same pooled items — a "
         "robust companion to the mean, and the number the older per-item "
         "reports used. <metric>_plate_spread_sd is the SD ACROSS the plate "
         "means, which is how this workbook used to compute its error bar. It "
         "is kept so a rogue dish stays visible, and only for that: it is a QC "
         "column, never an error bar."),
        ("Blank is not zero",
         "A blank cell means the quantity was not measured — no items, or none "
         "finite. Zero would be a claim about the plate. Non-finite values are "
         "dropped per column, so n_<metric> on per_plate can differ between "
         "columns of the same row."),
        ("What was excluded", keep_note),
        ("Units", unit_note),
        ("", ""),
        ("METRICS", ""),
    ]
    for m in metrics:
        rows.append((m.label + (f" ({m.unit})" if m.unit else ""),
                     (m.note or "") + (f"  [column: {m.key}, plate summary: "
                                       f"{m.agg}]")))
    return rows


def write_aggregate_sheets(wb, agg: AC.Aggregation) -> None:
    """plate_summary, condition_summary and qc, from one Aggregation.

    Deliberately NOT called per_plate / per_condition: counting already has
    sheets with those names and crawling already has a per_condition that pools
    worms rather than plates. New names mean nothing existing changes meaning
    under a reader who has scripts pointed at it, and the README says how the
    two relate.
    """
    metrics = list(agg.metrics)
    # A timecourse row is identified by (timepoint, condition) — the condition
    # alone repeats once per imaging day. Without this column the sheet holds
    # five indistinguishable copies of every condition, which is not a sheet
    # missing a nicety, it is a sheet nobody can read a number off.
    tc = bool(getattr(agg, "timepoints", []))
    tp_h = ["timepoint_h"] if tc else []

    def _tp(r):
        return [r.get("timepoint_h")] if tc else []

    hdr = tp_h + ["condition", "strain", "dose", "dose_unit", "plate",
                  "n_items", "n_kept"]
    for m in metrics:
        hdr += [m.key, f"n_{m.key}"]
    rows = []
    for p in agg.per_plate:
        row = _tp(p) + [p["condition"], p["strain"], p["dose"], p["unit"],
                        p["plate"], p["n_items"], p["n_kept"]]
        for m in metrics:
            row += [_round(p.get(m.key)), p.get(f"n_{m.key}")]
        rows.append(row)
    _sheet(wb, "plate_summary", hdr, rows, {"condition": 22, "plate": 16})

    hdr = tp_h + ["condition", "strain", "dose", "dose_unit", "name_parsed",
                  "n_plates", "n_plates_with_data", "n_items", "n_kept"]
    for m in metrics:
        hdr += [f"{m.key}_mean", f"{m.key}_sd", f"{m.key}_sem",
                f"{m.key}_n", f"{m.key}_pooled_median",
                f"{m.key}_plate_spread_sd"]
    rows = []
    for c in agg.per_condition:
        row = _tp(c) + [c["condition"], c["strain"], c["dose"], c["unit"],
                        bool(c["parsed"]), c["n_plates"],
                        c["n_plates_with_data"], c["n_items"], c["n_kept"]]
        for m in metrics:
            row += [_round(c.get(f"{m.key}_mean")),
                    _round(c.get(f"{m.key}_sd")),
                    _round(c.get(f"{m.key}_sem")),
                    c.get(f"{m.key}_n"),
                    _round(c.get(f"{m.key}_pooled_median")),
                    _round(c.get(f"{m.key}_plate_spread_sd"))]
        rows.append(row)
    _sheet(wb, "condition_summary", hdr, rows, {"condition": 22})

    hdr = tp_h + ["condition", "strain", "dose", "n_plates",
                  "n_plates_with_data", "n_items", "n_kept", "kept_pct",
                  "items_per_plate_mean", "items_per_plate_min",
                  "items_per_plate_max", "pct_of_control"]
    rows = []
    for c in agg.per_condition:
        # Its own timepoint's plates, never every day's — the same collision
        # the timepoint column exists to undo.
        plates = [p for p in agg.per_plate
                  if p["condition"] == c["condition"]
                  and (not tc or p.get("timepoint_h") == c.get("timepoint_h"))]
        kept = [p["n_kept"] for p in plates]
        ctrl = _control_for(c, agg.per_condition, by_timepoint=tc)
        rows.append(_tp(c) + [
            c["condition"], c["strain"], c["dose"], c["n_plates"],
            c["n_plates_with_data"], c["n_items"], c["n_kept"],
            _round(100.0 * c["n_kept"] / c["n_items"]) if c["n_items"] else None,
            _round(AC.mean(kept)), min(kept) if kept else None,
            max(kept) if kept else None,
            _round(100.0 * c["n_kept"] / ctrl) if ctrl else None,
        ])
    _sheet(wb, "qc", hdr, rows, {"condition": 22})


def write_condition_csv(path, agg: AC.Aggregation) -> int:
    """condition_summary as a CSV beside the workbook, same columns.

    A new file rather than a rewrite of the pipeline's existing summary CSV:
    motility's summarises videos and counting's summarises with a different set
    of columns, and silently changing what a file called *_summary.csv contains
    is how someone's downstream script starts reporting the wrong number without
    erroring.
    """
    import csv as _csv
    metrics = list(agg.metrics)
    tc = bool(getattr(agg, "timepoints", []))
    hdr = (["timepoint_h"] if tc else []) + [
        "condition", "strain", "dose", "dose_unit", "name_parsed",
        "n_plates", "n_plates_with_data", "n_items", "n_kept"]
    for m in metrics:
        hdr += [f"{m.key}_mean", f"{m.key}_sd", f"{m.key}_sem",
                f"{m.key}_n", f"{m.key}_pooled_median",
                f"{m.key}_plate_spread_sd"]
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = _csv.writer(fh)
        w.writerow(hdr)
        for c in agg.per_condition:
            row = ([c.get("timepoint_h")] if tc else []) + [
                   c["condition"], c["strain"],
                   "" if c["dose"] is None else c["dose"], c["unit"],
                   bool(c["parsed"]), c["n_plates"], c["n_plates_with_data"],
                   c["n_items"], c["n_kept"]]
            for m in metrics:
                row += [_blank(c.get(f"{m.key}_mean")),
                        _blank(c.get(f"{m.key}_sd")),
                        _blank(c.get(f"{m.key}_sem")),
                        _blank(c.get(f"{m.key}_n")),
                        _blank(c.get(f"{m.key}_pooled_median")),
                        _blank(c.get(f"{m.key}_plate_spread_sd"))]
            w.writerow(row)
    return len(agg.per_condition)


def _blank(v):
    """None -> empty cell. Never 0.0 — see the README's "Blank is not zero"."""
    r = _round(v)
    return "" if r is None else r


def _control_for(c: dict, cond_rows: Sequence[dict],
                 by_timepoint: bool = False) -> Optional[int]:
    """Kept-item count of the LOWEST dose of the SAME strain.

    Its own control, never another strain's — the same rule the Development
    explorer uses for % of control. In a timecourse it is also never another
    DAY's control: plates run down over four days, and charging that decline to
    the treatment is the exact error the timepoint column exists to prevent.
    """
    same = [r for r in cond_rows
            if r["strain"] == c["strain"] and r.get("dose") is not None
            and (not by_timepoint
                 or r.get("timepoint_h") == c.get("timepoint_h"))]
    if not same:
        return None
    lo = min(same, key=lambda r: r["dose"])
    return lo["n_kept"] or None


def write_distribution_sheets(wb, dist: Optional[dict], label: str,
                              unit: str) -> None:
    """size_histogram / size_summary, in the shape Development writes them."""
    if not dist:
        return
    keys = list(dist["groups"].keys())
    edges = dist["bin_edges"]
    u = unit.replace("/", "_per_").replace(" ", "") or "value"
    hdr = [f"bin_lo_{u}", f"bin_hi_{u}", f"bin_mid_{u}"] + keys
    rows = []
    for i in range(len(edges) - 1):
        mid = ((edges[i] * edges[i + 1]) ** 0.5 if dist.get("log")
               else 0.5 * (edges[i] + edges[i + 1]))
        rows.append([round(edges[i], 4), round(edges[i + 1], 4), round(mid, 4)]
                    + [dist["groups"][k]["hist"][i] for k in keys])
    _sheet(wb, "dist_histogram", hdr, rows, {hdr[0]: 14, hdr[1]: 14, hdr[2]: 14})

    hdr = ["group", "metric", "unit", "n", "p10", "p25", "p50_median",
           "p75", "p90", "mean"]
    rows = [[k, label, unit, g["n"], g["p10"], g["p25"], g["p50"], g["p75"],
             g["p90"], g["mean"]]
            for k, g in dist["groups"].items()]
    _sheet(wb, "dist_summary", hdr, rows, {"group": 26, "metric": 24})


def _round(v, nd: int = 4):
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return v
    return round(f, nd) if f == f else None


def run_info_common(assay: str, out_dir, n_conditions: int, n_plates: int,
                    n_items: int, n_kept: int, extra: Sequence[tuple] = ()) -> list:
    rows = [
        ("timestamp", datetime.now().isoformat(timespec="seconds")),
        ("assay", assay),
        ("output_dir", str(out_dir)),
        ("replication_unit", "item (worm / colony / well) — condition_summary "
                             "is the mean pooled over items with ±1 SEM, and "
                             "<metric>_n is the n. Plates are not replicates"),
        ("n_conditions", n_conditions),
        ("n_plates", n_plates),
        ("n_items_measured", n_items),
        ("n_items_kept", n_kept),
        ("items_kept_pct", round(100.0 * n_kept / n_items, 2) if n_items else ""),
    ]
    rows.extend(extra)
    return rows
