"""Tests for the shared assay layer: aggregation, sheets, CSV, explorer.

Run from launcher/:  python test_assay_reports.py

Dependency-light — no pytest, same style as test_survival_scale.py. Figures are
written and checked for existence only; nothing here looks at pixels.
"""
from __future__ import annotations

import csv
import json
import re
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import assay_common as AC          # noqa: E402
import assay_reports as AR         # noqa: E402

FAILURES: list[str] = []


def check(cond, what):
    print(("  PASS  " if cond else "  FAIL  ") + what)
    if not cond:
        FAILURES.append(what)


def near(a, b, tol=1e-6):
    return a is not None and abs(float(a) - float(b)) <= tol


# ---------------------------------------------------------------------------

def test_grammar():
    print("\ncondition grammar")
    check(AC.parse_condition("601 20J") == ("601", 20, "J/m²"),
          "\"601 20J\" parses to strain, dose, unit")
    check(AC.parse_condition("N2 100 uM") == ("N2", 100, "µM"),
          "a spaced µM dose parses")
    check(AC.parse_condition("control") is None,
          "a name outside the grammar does not parse")
    f = AC.split_condition("control")
    check(f["strain"] == "control" and f["dose"] is None and not f["parsed"],
          "an unparsed name becomes its own strain rather than being dropped")
    import survival
    check(survival.parse_condition is AC.parse_condition,
          "Development and the other assays share ONE grammar object")


def test_plate_first():
    print("\nthe replication unit is the plate")
    m = [AC.Metric("bpm", "Bend rate", "bends/min")]
    items = ([{"condition": "601 20J", "plate": "p1", "bpm": 10.0, "keep": True}] * 100
             + [{"condition": "601 20J", "plate": "p2", "bpm": 100.0, "keep": True}] * 2)
    agg = AC.aggregate(items, m, keep=lambda r: r["keep"])
    c = agg.per_condition[0]
    check(near(c["bpm_mean"], 55.0),
          f"a 100-worm plate does not outvote a 2-worm one (mean {c['bpm_mean']})")
    check(near(c["bpm_pooled_median"], 10.0),
          "the pooled median is kept alongside, and differs (10.0)")
    check(c["n_plates"] == 2 and c["n_kept"] == 102,
          "n_plates is 2 and n_kept is 102 — both are reported")
    check(near(c["bpm_sd"], 63.6396, 1e-3),
          "SD is across the two plate means, not across worms")

    one = AC.aggregate([{"condition": "N2 0J", "plate": "p1", "bpm": 5.0}], m)
    check(one.per_condition[0]["bpm_sd"] is None,
          "a single plate has no SD — blank, not 0.0")


def test_gate():
    print("\nquality gate")
    m = [AC.Metric("bpm", "Bend rate", "bends/min")]
    items = [{"condition": "N2 0J", "plate": "p1", "bpm": 10.0, "is_long": True},
             {"condition": "N2 0J", "plate": "p1", "bpm": 90.0, "is_long": False}]
    agg = AC.aggregate(items, m, keep=lambda r: r["is_long"])
    p = agg.per_plate[0]
    check(p["n_items"] == 2 and p["n_kept"] == 1,
          "a failing item is counted in n_items and excluded from n_kept")
    check(near(p["bpm"], 10.0), "…and excluded from the statistic")

    items = [{"condition": "N2 0J", "plate": "p1", "bpm": None},
             {"condition": "N2 0J", "plate": "p1", "bpm": float("nan")}]
    agg = AC.aggregate(items, m)
    check(agg.per_plate[0]["bpm"] is None and agg.per_plate[0]["n_bpm"] == 0,
          "all-non-finite gives blank and n=0, never 0.0")


def _worms(n_plates=3, n_worms=12):
    rows = []
    for strain, base in (("601", 40.0), ("N2", 90.0)):
        for dose in (0, 20, 40):
            for p in range(n_plates):
                for w in range(n_worms):
                    hit = base - (dose * 0.6 if strain == "601" else 0.05 * dose)
                    rows.append({
                        "condition": f"{strain} {dose}J", "plate": f"plate {p+1:02d}",
                        "bpm": max(0.0, hit + (w % 5) - 2 + p),
                        "bend_interval_cv": 0.2 + 0.01 * (w % 7),
                        "speed_median_abs": 30.0 + w,
                        "duration_s": 25.0 + w,
                        "is_long": w != 0,          # one short worm per plate
                        "mean_speed_pxs": 12.0 + w * 0.5,
                        # The crawling headline is BL/s, not px/s. A fixture
                        # that carries only the pixel column makes the report
                        # skip its distribution panel and the test then checks
                        # a None — which is how this fixture last drifted.
                        "mean_speed_bls": (12.0 + w * 0.5) / 100.0,
                        "fraction_paused": 0.1 + 0.01 * (w % 5),
                        "is_immobile": 1.0 if w < 2 else 0.0,
                        "reversal_rate_per_min": 2.0 + 0.1 * w,
                        "reversal_rate_moving_per_min": (2.0 + 0.1 * w) / 0.9,
                        "turn_rate_per_min": 1.0 + 0.05 * w,
                        "tortuosity": 1.2 + 0.02 * w,
                        "directionality": 0.8 - 0.02 * w,
                        "net_displacement_bl": 3.0 + 0.1 * w,
                        "passed_filter": w != 0,
                    })
    return rows


def _timecourse_worms(hours=(0.0, 24.0, 48.0)):
    """The same worms imaged on several days, with a real decline in the
    treated arm, so a test can tell a timepoint-aware output from one that
    silently kept whichever day came last."""
    rows = []
    for i, t in enumerate(hours):
        for r in _worms(n_plates=2, n_worms=8):
            r = dict(r)
            r["timepoint_h"] = t
            r["source_folder"] = f"day{i}"
            if r["condition"].endswith("40J"):
                r["mean_speed_bls"] *= (1.0 - 0.3 * i)
            rows.append(r)
    return rows


def _run(fn, tmp: Path, *args, **kw):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    logs: list[str] = []
    out = fn(wb, *args, out_dir=tmp, write_log=logs.append, **kw) \
        if False else fn(wb, *args, tmp, logs.append, **kw)
    ws = wb.create_sheet("_keepalive") if not wb.sheetnames else None
    wb.save(tmp / "book.xlsx")
    return wb, out, logs


def test_motility(tmp: Path):
    print("\nmotility report")
    d = tmp / "mot"; d.mkdir()
    wb, written, logs = _run(AR.motility_report, d, _worms(), long_threshold_s=5.0)
    names = wb.sheetnames
    check("README" in names and "run_info" in names, "README and run_info written")
    check("plate_summary" in names and "condition_summary" in names
          and "qc" in names, "plate_summary, condition_summary and qc written")
    check("dist_histogram" in names and "dist_summary" in names,
          "the distribution sheets are written")
    check((d / "motility_condition_summary.csv").exists(),
          "a condition-summary CSV is written beside the workbook")
    check((d / "motility_dose_response.png").exists(),
          "the dose-response figure is written")
    check((d / "motility_distribution.png").exists(),
          "the distribution figure is written")
    check((d / "explorer.html").exists(), "explorer.html is written")

    with open(d / "motility_condition_summary.csv", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    check(len(rows) == 6, f"one CSV row per condition (got {len(rows)})")
    check(rows[0]["n_plates"] == "3", "n_plates is on every row")
    c601_0 = next(r for r in rows if r["condition"] == "601 0J")
    c601_40 = next(r for r in rows if r["condition"] == "601 40J")
    check(float(c601_0["bpm_mean"]) > float(c601_40["bpm_mean"]),
          "the seeded dose effect survives aggregation")
    check(rows[0]["strain"] and rows[0]["dose"], "strain and dose are parsed out")

    html = (d / "explorer.html").read_text(encoding="utf-8")
    blob = re.search(r"const D = (\{.*?\});\n", html, re.S)
    check(blob is not None, "the payload is inlined in the explorer")
    if blob:
        D = json.loads(blob.group(1))
        check(D["has_dose"] and D["doses"] == [0, 20, 40],
              "the explorer knows its dose axis")
        check(len(D["cond"]) == 6 and len(D["plates"]) == 18,
              "every condition and plate is in the payload")
        check(D["dist"] and D["dist"]["log"] is False,
              "bend rate is drawn in linear space, not log")
        check(all("|" in k for k in D["dist"]["groups"]),
              "distribution groups are keyed strain|dose")
    check(any("explorer.html" in m for m in logs), "the explorer write is logged")


def test_crawling(tmp: Path):
    print("\ncrawling report")
    d = tmp / "crawl"; d.mkdir()
    wb, written, logs = _run(AR.crawling_report, d, _worms(), min_span_s=30.0)
    check("condition_summary" in wb.sheetnames,
          "condition_summary is added without touching per_worm/per_condition")
    check((d / "crawling_condition_summary.csv").exists()
          and (d / "explorer.html").exists(),
          "CSV and explorer are written")
    html = (d / "explorer.html").read_text(encoding="utf-8")
    D = json.loads(re.search(r"const D = (\{.*?\});\n", html, re.S).group(1))
    check([m["key"] for m in D["metrics"]] ==
          ["mean_speed_bls", "fraction_paused", "is_immobile", "directionality",
           "mean_speed_pxs", "reversal_rate_moving_per_min",
           "reversal_rate_per_min", "turn_rate_per_min", "net_displacement_bl",
           "bpm", "tortuosity"],
          "the agreed crawling panels, in order")
    check([m["key"] for m in D["metrics"] if m["headline"]] ==
          ["mean_speed_bls", "fraction_paused", "is_immobile", "directionality"],
          "the four that take a row in the figures are flagged headline")
    check(D["dist"]["log"] is True, "speed is drawn in log space")
    check(D["dist"]["unit"] == "BL/s",
          "the distribution is the body-length speed, not the pixel one")
    check(not (D.get("caveat") or "").strip(),
          "no standing caveat banner — the workbook README carries that now")
    check(not any("motility" in (m.get("note") or "").lower()
                  for m in D["metrics"]),
          "no metric note sends the reader to another assay")


def test_timecourse(tmp: Path):
    """The regression this whole change exists for.

    Before it, a multi-folder run aggregated by timepoint and then every
    consumer of that aggregation dropped the timepoint again: the sheets held
    N indistinguishable copies of each condition, the dose-response figure drew
    one arbitrary day as the condition mean with every day's plates behind it,
    and the explorer received rows it could not tell apart. Each of those is a
    silent wrong answer, so each gets an assertion.
    """
    print("\ncrawling timecourse")
    d = tmp / "tc"; d.mkdir()
    hours = (0.0, 24.0, 48.0)
    wb, written, logs = _run(AR.crawling_report, d, _timecourse_worms(hours),
                             min_span_s=30.0, by_timepoint=True)

    # plate_summary is one row per (timepoint, condition, PLATE); the other two
    # are one row per (timepoint, condition). Checking the wrong key would pass
    # for the wrong reason, so each sheet is checked against its own.
    for sheet, key in (("plate_summary", ("timepoint_h", "condition", "plate")),
                       ("condition_summary", ("timepoint_h", "condition")),
                       ("qc", ("timepoint_h", "condition"))):
        hdr = [c.value for c in wb[sheet][1]]
        check(hdr[0] == "timepoint_h", f"{sheet} leads with timepoint_h")
        ix = [hdr.index(k) for k in key]
        seen = {tuple(r[i].value for i in ix)
                for r in wb[sheet].iter_rows(min_row=2)}
        check(len(seen) == wb[sheet].max_row - 1,
              f"{sheet} has no two rows sharing {key}")

    with open(d / "crawling_condition_summary.csv", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    check(rows and "timepoint_h" in rows[0],
          "the condition CSV carries the timepoint too")
    check(len({r["condition"] for r in rows}) * len(hours) == len(rows),
          "one CSV row per condition per timepoint, none collapsed")

    check((d / "crawling_timecourse.png").exists(),
          "the timecourse figure is written")
    check((d / "crawling_normalised.png").exists(),
          "the same-day-control figure is written")

    html = (d / "explorer.html").read_text(encoding="utf-8")
    D = json.loads(re.search(r"const D = (\{.*?\});\n", html, re.S).group(1))
    check(D["timepoints"] == list(hours), "the explorer knows its timepoints")
    check(all(c.get("tp") is not None for c in D["cond"]),
          "every condition row in the payload carries its timepoint")
    check(all(p.get("tp") is not None for p in D["plates"]),
          "so does every plate row")
    check(len({(c["condition"], c["tp"]) for c in D["cond"]}) == len(D["cond"]),
          "no two payload rows are indistinguishable")
    keys = list((D["dist"] or {}).get("group_meta", {}))
    check(keys and all("@" in k for k in keys),
          "distribution groups are split by timepoint, not pooled across days")
    norm = D.get("normalised")
    check(norm is not None, "the payload carries the same-day-control series")
    dosed = [s for s in (norm or {}).get("series", []) if s["strain"] == "601"
             and s["dose"] == 40]
    check(bool(dosed) and dosed[0]["t_half"] is not None,
          "a condition that halves against its own control reports a t_half")

    # The single-folder path must be untouched by all of this.
    d2 = tmp / "tc_single"; d2.mkdir()
    wb2, _, _ = _run(AR.crawling_report, d2, _worms(), min_span_s=30.0)
    check([c.value for c in wb2["condition_summary"][1]][0] == "condition",
          "a single-folder run gets no timepoint column")
    check(not (d2 / "crawling_timecourse.png").exists(),
          "…and no timecourse figure")


def test_counting(tmp: Path):
    print("\ncounting report")
    d = tmp / "count"; d.mkdir()
    plates, colonies = [], []
    for dose in (0, 2, 4):
        for p in range(3):
            n = max(1, 60 - 14 * dose)
            plates.append({"condition": f"HeLa {dose}J", "plate": f"well {p+1}",
                           "colony_count": n + p, "mean_area_mm2": 0.8,
                           "median_area_mm2": 0.7,
                           "total_colony_area_mm2": 0.8 * (n + p),
                           "stained_fraction": 0.05 * (dose + 1),
                           "confluent": dose == 0 and p == 0})
            for k in range(n):
                colonies.append({"condition": f"HeLa {dose}J",
                                 "plate": f"well {p+1}",
                                 "equiv_diam_um": 300.0 + 8 * (k % 20)})
    wb, written, logs = _run(AR.counting_report, d, plates, colonies)
    check("condition_summary" in wb.sheetnames and "qc" in wb.sheetnames,
          "condition_summary and qc are added")
    check((d / "counting_condition_summary.csv").exists()
          and (d / "explorer.html").exists()
          and (d / "counting_dose_response.png").exists(),
          "CSV, explorer and figure are written")
    html = (d / "explorer.html").read_text(encoding="utf-8")
    D = json.loads(re.search(r"const D = (\{.*?\});\n", html, re.S).group(1))
    lo = next(c for c in D["cond"] if c["dose"] == 0)
    hi = next(c for c in D["cond"] if c["dose"] == 4)
    check(lo["stats"]["colony_count"]["mean"] > hi["stats"]["colony_count"]["mean"],
          "colony count falls with dose")
    check(lo["n_plates"] == 3, "each well is one plate row")
    check("confluent" in (D.get("caveat") or "").lower(),
          "the confluent wells are called out on the page")
    check(D["dist"]["log"] is True and D["dist"]["unit"] == "µm",
          "colony size is a log-space distribution in µm")


def test_logger_cannot_kill_a_run(tmp: Path):
    """Regression: a write_log that raises must not cost the run its outputs.

    Motility and crawling build log.txt inside a `with open(...)` block that has
    already closed by the time the workbook is written, so the write_log closure
    they hand us was writing to a closed file and raising ValueError. That
    aborted the report after the CSV and before the figures, and the handler
    raised again logging the failure, which escaped and took the pipeline's own
    summary CSV and overview figure with it. Seen in the wild 2026-08-19.
    """
    print("\nlogging failures cannot end a run")
    from openpyxl import Workbook

    def closed_file_logger(msg):
        raise ValueError("I/O operation on closed file.")

    d = tmp / "closedlog"; d.mkdir()
    wb = Workbook(); wb.remove(wb.active)
    AR.motility_report(wb, _worms(), d, closed_file_logger, long_threshold_s=5.0)
    check("condition_summary" in wb.sheetnames,
          "the workbook sheets are still written")
    check((d / "motility_condition_summary.csv").exists(),
          "the condition CSV is still written")
    check((d / "motility_dose_response.png").exists()
          and (d / "motility_distribution.png").exists(),
          "BOTH figures are still written (these were the casualties)")
    check((d / "explorer.html").exists(),
          "the explorer is still written")

    d2 = tmp / "nolog"; d2.mkdir()
    wb2 = Workbook(); wb2.remove(wb2.active)
    AR.motility_report(wb2, _worms(), d2, None, long_threshold_s=5.0)
    check((d2 / "explorer.html").exists(),
          "write_log=None is also survivable")


if __name__ == "__main__":
    with tempfile.TemporaryDirectory() as t:
        tmp = Path(t)
        test_grammar()
        test_plate_first()
        test_gate()
        test_motility(tmp)
        test_crawling(tmp)
        test_timecourse(tmp)
        test_counting(tmp)
        test_logger_cannot_kill_a_run(tmp)
    print()
    if FAILURES:
        print(f"{len(FAILURES)} FAILURE(S):")
        for f in FAILURES:
            print("  - " + f)
        sys.exit(1)
    print("all checks passed")
