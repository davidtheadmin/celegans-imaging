"""
Development workbook (development_results.xlsx).

Everything the four figures draw is in here, as LIVE FORMULAS wherever the
inputs are also in the workbook. Only three kinds of cell hold a literal:

  * per-image stage counts — the raw measurement, nothing to derive them from;
  * the body-size sheets — derived from soft_stage_scores.csv, whose 5,000 to
    17,000 per-animal rows deliberately do NOT go into the workbook;
  * error / gap counts, which describe what did not happen.

Everything else — totals, stage index, composition %, survival %, the QC
numbers — is a formula, so editing a count makes the whole sheet move.

How the formulas stay simple
----------------------------
per_image and per_plate are both sorted by (timepoint, strain, dose, plate),
which makes the rows belonging to one plate, and the plates belonging to one
condition, CONTIGUOUS. A derived cell is then a plain SUM/AVERAGE/STDEV over an
exact row range rather than a SUMIFS whose criteria have to be trusted. The
ranges are computed once, in Python, from the same grouping the rows were
written in.

That is also the failure mode to watch: an off-by-one range recalculates
perfectly and reports the wrong number. ``verify_workbook`` re-reads the
recalculated file and compares EVERY computed cell against the value Python
already has. Do not delete it.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

# openpyxl / pandas are heavy; imported inside the functions so importing this
# module at launcher start-up stays cheap (same rule as survival.py).

_FONT = "Arial"

# Number formats
_F_INT = "#,##0"
_F_1 = "0.0"
_F_2 = "0.00"
_F_PCT = "0.0"


def _col(idx: int) -> str:
    """1 -> A, 27 -> AA."""
    s = ""
    while idx:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def _sheet_ref(sheet: str, col: str, r1: int, r2: int) -> str:
    """Quoted cross-sheet range. Sheet names here have no spaces, but quoting
    is free and an unquoted name with a space silently evaluates to #VALUE!."""
    return f"'{sheet}'!${col}${r1}:${col}${r2}"


class _Grid:
    """Tracks where each logical column ended up, so formulas never guess."""

    def __init__(self, headers: list[str]) -> None:
        self.headers = list(headers)
        self.pos = {h: i + 1 for i, h in enumerate(headers)}

    def col(self, name: str) -> str:
        return _col(self.pos[name])


def _write_header(ws, headers: list[str]) -> None:
    from openpyxl.styles import Alignment, Font
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=_FONT, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="bottom",
                                   wrap_text=True)
    ws.freeze_panes = "A2"


def _style_body(ws, n_rows: int, n_cols: int) -> None:
    from openpyxl.styles import Font
    f = Font(name=_FONT, size=10)
    for r in range(2, n_rows + 2):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).font = f


def _autosize(ws, headers: list[str], widths: Optional[dict] = None) -> None:
    for i, h in enumerate(headers, start=1):
        w = (widths or {}).get(h)
        ws.column_dimensions[_col(i)].width = w if w else max(9, min(22, len(h) + 3))


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _readme_rows(meta: dict, plans, summary: dict, stage_names: list[str],
                 unmapped: list[str], size_note: str) -> list[tuple[str, str]]:
    rescore = meta.get("rescore") or {}
    alpha = float(rescore.get("alpha") or 0.0)
    rows: list[tuple[str, str]] = [
        ("WHAT THIS WORKBOOK IS", ""),
        ("", "A Development run: the YOLO staging model was run over every "
             "image in the folder(s) below, each animal was assigned a "
             "developmental stage, and those calls were aggregated per image, "
             "per plate and per condition."),
        ("", ""),
        ("THE READOUT", ""),
        ("Mean stage index",
         "L1=1, L2=2, L3=3, L4=4, adult=5. Eggs are EXCLUDED — an egg has no "
         "position on this scale, and counting it as a 0 would drag the mean "
         "by however many eggs happened to be laid. The mean is computed per "
         "plate first, then averaged across plates, so a 600-animal plate does "
         "not outvote a 25-animal one."),
        ("Replication unit",
         "Plates, when a condition has more than one. When a condition has a "
         "single plate the four quadrant images of that plate are used instead, "
         "because otherwise there is no spread to report at all. Every "
         "per_condition row says which it used, in 'replicate_unit'. A mean ± SD "
         "whose n means two different things in two rows is worse than none."),
        ("Body size",
         "sqrt(width x height) of the detection box. Reported in MICROMETRES "
         "when every image carries a spatial calibration in its TIFF tags, and "
         "in full-frame pixels otherwise — all or nothing, never a mixture; "
         "run_info records which and why. Distributions are estimated in LOG "
         "space, because growth is multiplicative: a fixed bandwidth in linear "
         "space over-smooths the tight L1 peak and under-smooths the broad "
         "adult tail."),
        ("...and what it is not",
         "APPARENT size, not a body length. A box does not know how coiled the "
         "animal in it is, so a curled worm reads smaller than a straight one "
         "of the same length: sqrt(w x h) grows about 2.5x from L1 to adult "
         "where real body length grows 4-5x. Micrometres make it comparable "
         "across magnifications; they do not make it a length, and it should "
         "not be compared against published body lengths."),
        ("", ""),
        ("SURVIVAL % — READ THIS BEFORE USING IT", ""),
        ("", "survival % = (L3 + L4 + adult) / (all staged animals) x 100. It is "
             "in per_image, per_plate and per_condition because people ask for "
             "it, and dropping it silently would be worse than keeping it with a "
             "warning. It is in NO figure, deliberately, for two reasons:"),
        ("", "1. The denominator collapses. In a full dose experiment one strain "
             "lost 95% of its animals by 20 J/m², so survival % was computed "
             "over ~25 surviving worms per plate and ROSE with dose — an "
             "inverted dose response that is an artefact of the shrinking "
             "denominator, not biology."),
        ("", "2. The survivor cutoff sits exactly on the L2/L3 boundary, which "
             "is where the model is weakest: about 26% of L3 calls are L2-sized. "
             "The cutoff is placed on the least reliable distinction available."),
        ("", "Use stage index, composition and body size for the claim; use "
             "survival % only when someone specifically asks for that number, "
             "and quote the caveat with it."),
        ("", ""),
        ("CLASS-CONFIDENCE CORRECTION (rescoring)", ""),
        ("Status", (f"ON, alpha {alpha:g}" if alpha
                    else "OFF (alpha 0 — arg-max on raw model scores)")),
        ("What it does",
         "The stage classes are not scored on a common scale: measured over "
         "17,084 detections, the L2 head's 99th-percentile score is 0.14 while "
         "L1 and L3 routinely reach 0.80. An arg-max across raw scores therefore "
         "compares numbers that are not comparable, and L2 loses nearly every "
         "contest it enters. The pass divides each class score by its reference "
         "raised to alpha, then re-picks the winner."),
        ("What it does not do",
         "It RELABELS ONLY. It runs after every suppression step, so the number "
         "of animals found and every box coordinate are unchanged at any alpha. "
         "A shift in these numbers after turning it on is pure "
         "reclassification — check that the totals match, and they will."),
        ("Caveat — not calibration",
         "Dividing by a per-class constant implicitly assumes the classes are "
         "equally common, and they are not: adult is ~44% of detections and L2 "
         "~0.4%. alpha is a lever on a prior, not a calibration. Validate "
         "against hand-labelled animals before any rescored count is reported "
         "as data."),
        ("Why alpha 2.0",
         "David compared several alphas against his own manual counts and found "
         "1.5-2.0 matched best, and chose 2.0. Recorded counter-evidence: at "
         "alpha 2 about 19% of newly-created L2 calls fall outside the plausible "
         "L2 size window (10% at alpha 1, 15% at 1.5), and the per-plate "
         "difference between 1.5 and 2.0 (mean 1.6 points) is smaller than the "
         "plate-to-plate noise in most conditions, so that comparison probably "
         "cannot separate them. The value lives in "
         "launcher/vision/stage_conf.json and is tunable without a rebuild."),
        ("", ""),
        ("KNOWN LIMITS", ""),
        ("L2/L3 boundary",
         "Unreliable, and not fixed here. There is no ground truth to place it: "
         "reassigning on a guessed size cutoff moved per-plate survival by up to "
         "42 points in test data. The likely cause is not class imbalance — "
         "posture moves a bounding box by ~15% while the L1->L2 and L2->L3 "
         "biological steps are only 15-17%, so the measurement barely resolves "
         "the boundary. Awaiting hand labels."),
        ("Counts, not rates",
         "Count-based relative survival (adults per plate, normalised to the "
         "condition's own control) is the readout that behaves when the "
         "denominator collapses. It is not part of this build; the qc sheet "
         "carries the raw material for it (animals per plate, % of control)."),
        ("", ""),
        ("WHICH CELLS ARE LIVE", ""),
        ("Formulas",
         "Every total, stage index, composition %, survival % and QC number is "
         "a formula over the sheets in this workbook. Edit a per_image count and "
         "everything downstream moves."),
        ("Literals",
         "per_image stage counts (the raw measurement); the two size sheets "
         "(derived from soft_stage_scores.csv, whose per-animal rows are "
         "deliberately not in this workbook); and error/gap counts."),
        ("Blank in a script?",
         "Formula cells are written without a cached result, so Excel fills "
         "them the first moment you open the file. Until then a script reading "
         "this workbook (pandas, openpyxl data_only=True) sees blanks where "
         "the formulas are. Open it in Excel and save once, or read the raw "
         "counts and recompute. This is normal, not a corrupt file."),
        ("Body-size provenance", size_note),
        ("", ""),
        ("REUSED WORK", ""),
        ("What was reused",
         f"{summary.get('n_reused', 0)} image(s) were not analysed again: "
         "their detections were read back from an earlier run of the same "
         f"folder. {summary.get('n_analysed', 0)} image(s) went through the "
         "model this time."
         + ("  The model was not run at all for this workbook."
            if summary.get("from_cache_only") else "")),
        ("Why that is safe",
         "Reuse is decided per image, on file size and modification time, and "
         "only against a run whose detection settings and model file are "
         "identical to this one's — the per-class confidence floors and the "
         "size gate are applied to candidate boxes BEFORE the merge, so "
         "changing either would change which boxes exist and invalidates the "
         "cache automatically."),
        ("The one thing that is recomputed",
         "The class-confidence correction. It is the last pass and a pure "
         "arg-max over the per-class scores, which are saved per detection, so "
         "switching it on or off — or changing alpha — is recalculated from "
         "the saved rows and gives exactly what a fresh run would have given. "
         f"{summary.get('n_relabelled', 0)} detection(s) were relabelled that "
         "way for this workbook."),
        ("", ""),
        ("PROVENANCE", ""),
        ("Written", datetime.now().isoformat(timespec="seconds")),
        ("Model", str(meta.get("_model_path") or "")),
        ("Model classes", ", ".join(stage_names)),
        ("Unmapped stages",
         ", ".join(unmapped) if unmapped else
         "(none — every class the model reported has a survivor mapping)"),
        ("Folders in this run", str(len(plans))),
    ]
    for p in plans:
        rows.append((f"  {p.folder.name}",
                     f"timepoint {p.hours:g} h — {p.detail}"))
    return rows


def _write_readme(wb, rows: list[tuple[str, str]]) -> None:
    from openpyxl.styles import Alignment, Font
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 110
    for i, (k, v) in enumerate(rows, start=1):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v)
        heading = k and not v
        a.font = Font(name=_FONT, size=10, bold=bool(heading or (k and v)))
        b.font = Font(name=_FONT, size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        a.alignment = Alignment(vertical="top")


def _write_run_info(wb, rows: list[tuple[str, object]]) -> None:
    from openpyxl.styles import Alignment, Font
    ws = wb.create_sheet("run_info")
    _write_header(ws, ["key", "value"])
    for k, v in rows:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 110
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(name=_FONT, size=10, bold=True)
        c = ws.cell(row=r, column=2)
        c.font = Font(name=_FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# The three count sheets, with their formula scaffolding
# ---------------------------------------------------------------------------

def _survivor_stages(stage_cols: list[str], cats: dict) -> tuple[list[str], list[str]]:
    surv = [s for s in stage_cols if cats.get(s) == "survivor"]
    non = [s for s in stage_cols if cats.get(s) == "non_survivor"]
    return surv, non


def _sum_expr(grid: _Grid, names: list[str], row: int) -> str:
    """A+B+C over named columns of one row; '0' when the list is empty."""
    if not names:
        return "0"
    return "+".join(f"{grid.col(n)}{row}" for n in names)


def _weighted_expr(grid: _Grid, weights: list[tuple[str, float]], row: int) -> str:
    if not weights:
        return "0"
    return "+".join(f"{w:g}*{grid.col(n)}{row}" for n, w in weights)


def _write_counts_sheet(wb, name: str, rows: list[dict], grid: _Grid,
                        formula_builder: Callable[[int, dict], dict]):
    """Write one of per_image / per_plate. formula_builder returns
    {header: formula_string} for the derived columns of that row."""
    ws = wb.create_sheet(name)
    _write_header(ws, grid.headers)
    for i, r in enumerate(rows):
        excel_row = i + 2
        formulas = formula_builder(excel_row, r)
        out = []
        for h in grid.headers:
            if h in formulas:
                out.append(formulas[h])
            else:
                v = r.get(h)
                out.append("" if v is None else v)
        ws.append(out)
    _style_body(ws, len(rows), len(grid.headers))
    _autosize(ws, grid.headers, {"image": 30, "condition": 16, "folder": 22})
    return ws


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def write_workbook(out_path: Path, agg: dict, *, stage_names: list[str],
                   cats: dict, unmapped: list[str], meta: dict, plans,
                   summary: dict, size: Optional[dict],
                   write_log: Callable[[str], None]) -> dict:
    """Build development_results.xlsx. Returns a plan of every computed cell
    (address -> expected value) for verify_workbook to check after recalc."""
    from openpyxl import Workbook

    stage_cols = list(agg["stage_cols"])
    surv_stages, non_stages = _survivor_stages(stage_cols, cats)
    # Stages that carry a developmental index (everything but eggs / unmapped).
    indexed: list[tuple[str, float]] = []
    for st in stage_cols:
        w = index_map_lookup(st)
        if w is not None:
            indexed.append((st, w))
    indexed_names = [s for s, _ in indexed]

    wb = Workbook()
    wb.remove(wb.active)

    size_note = (
        f"{size['n_total']:,} animals from soft_stage_scores.csv, "
        f"{len(size['groups'])} condition x timepoint group(s). Sizes are in "
        f"{'micrometres' if size.get('unit') == 'um' else 'pixels'} "
        f"({size.get('scale_note', '')}). This is APPARENT size, sqrt(w*h) of "
        "the detection box, not a body length: a coiled animal reads smaller "
        "than a straight one of the same length."
        if size else
        "soft_stage_scores.csv held no usable size_px column — the two size "
        "sheets and the body-size figure were skipped."
    )
    _write_readme(wb, _readme_rows(meta, plans, summary, stage_names, unmapped,
                                   size_note))
    _write_run_info(wb, _run_info_rows(meta, plans, summary, stage_names,
                                       unmapped, agg, size))

    expected: dict[str, float] = {}

    # ---- per_image --------------------------------------------------------
    img_rows = agg["per_image"]
    img_headers = (
        ["folder", "timepoint_h", "condition", "strain", "dose", "unit",
         "plate", "quadrant", "image"] + stage_cols
        + ["unmapped", "total", "n_staged", "stage_index",
           "n_survivors", "n_non_survivors", "survival_pct"]
    )
    ig = _Grid(img_headers)
    first_stage, last_stage = ig.col(stage_cols[0]), ig.col(stage_cols[-1])

    def img_formulas(r: int, row: dict) -> dict:
        n_staged = f"{ig.col('n_staged')}{r}"
        surv = f"{ig.col('n_survivors')}{r}"
        non = f"{ig.col('n_non_survivors')}{r}"
        return {
            "total": f"=SUM({first_stage}{r}:{last_stage}{r})",
            "n_staged": f"={_sum_expr(ig, indexed_names, r)}",
            "stage_index": (f"=IF({n_staged}=0,\"\","
                            f"({_weighted_expr(ig, indexed, r)})/{n_staged})"),
            "n_survivors": f"={_sum_expr(ig, surv_stages, r)}",
            "n_non_survivors": f"={_sum_expr(ig, non_stages, r)}",
            "survival_pct": (f"=IF({surv}+{non}=0,\"\","
                             f"100*{surv}/({surv}+{non}))"),
        }

    _write_counts_sheet(wb, "per_image", img_rows, ig, img_formulas)
    for i, row in enumerate(img_rows):
        r = i + 2
        for h in ("total", "n_staged", "stage_index", "n_survivors",
                  "n_non_survivors", "survival_pct"):
            expected[f"per_image!{ig.col(h)}{r}"] = row[h]

    # Row ranges: which per_image rows belong to which plate.
    img_span: dict[tuple, tuple[int, int]] = {}
    for i, row in enumerate(img_rows):
        key = (row["timepoint_h"], row["condition"], row["plate"])
        r = i + 2
        lo, hi = img_span.get(key, (r, r))
        img_span[key] = (min(lo, r), max(hi, r))

    # ---- per_plate --------------------------------------------------------
    plate_rows = agg["per_plate"]
    plate_headers = (
        ["folder", "timepoint_h", "condition", "strain", "dose", "unit", "plate"]
        + stage_cols
        + ["unmapped", "total", "n_staged", "stage_index",
           "n_survivors", "n_non_survivors", "survival_pct",
           "n_images", "quadrant_cv_pct"]
    )
    pg = _Grid(plate_headers)
    p_first, p_last = pg.col(stage_cols[0]), pg.col(stage_cols[-1])

    def plate_formulas(r: int, row: dict) -> dict:
        key = (row["timepoint_h"], row["condition"], row["plate"])
        lo, hi = img_span[key]
        n_staged = f"{pg.col('n_staged')}{r}"
        surv = f"{pg.col('n_survivors')}{r}"
        non = f"{pg.col('n_non_survivors')}{r}"
        tot_rng = _sheet_ref("per_image", ig.col("total"), lo, hi)
        f = {
            "total": f"=SUM({p_first}{r}:{p_last}{r})",
            "n_staged": f"={_sum_expr(pg, indexed_names, r)}",
            "stage_index": (f"=IF({n_staged}=0,\"\","
                            f"({_weighted_expr(pg, indexed, r)})/{n_staged})"),
            "n_survivors": f"={_sum_expr(pg, surv_stages, r)}",
            "n_non_survivors": f"={_sum_expr(pg, non_stages, r)}",
            "survival_pct": (f"=IF({surv}+{non}=0,\"\","
                             f"100*{surv}/({surv}+{non}))"),
            "n_images": f"=COUNT({tot_rng})",
            # quadrant-to-quadrant spread within this plate, as a % of its own
            # mean. Needs >=2 images and a non-zero mean, hence the guard.
            "quadrant_cv_pct": (
                f"=IF(OR(COUNT({tot_rng})<2,AVERAGE({tot_rng})=0),\"\","
                f"100*STDEV({tot_rng})/AVERAGE({tot_rng}))"),
        }
        # Stage counts summed straight from the images of this plate.
        for st in stage_cols:
            f[st] = f"=SUM({_sheet_ref('per_image', ig.col(st), lo, hi)})"
        f["unmapped"] = f"=SUM({_sheet_ref('per_image', ig.col('unmapped'), lo, hi)})"
        return f

    _write_counts_sheet(wb, "per_plate", plate_rows, pg, plate_formulas)
    for i, row in enumerate(plate_rows):
        r = i + 2
        for h in stage_cols + ["unmapped", "total", "n_staged", "stage_index",
                               "n_survivors", "n_non_survivors",
                               "survival_pct", "n_images"]:
            expected[f"per_plate!{pg.col(h)}{r}"] = row[h]

    plate_span: dict[tuple, tuple[int, int]] = {}
    for i, row in enumerate(plate_rows):
        key = (row["timepoint_h"], row["condition"])
        r = i + 2
        lo, hi = plate_span.get(key, (r, r))
        plate_span[key] = (min(lo, r), max(hi, r))
    cond_img_span: dict[tuple, tuple[int, int]] = {}
    for i, row in enumerate(img_rows):
        key = (row["timepoint_h"], row["condition"])
        r = i + 2
        lo, hi = cond_img_span.get(key, (r, r))
        cond_img_span[key] = (min(lo, r), max(hi, r))

    # ---- per_condition ----------------------------------------------------
    cond_rows = agg["per_condition"]
    cond_headers = (
        ["timepoint_h", "condition", "strain", "dose", "unit",
         "n_plates", "n_images", "replicate_unit", "n_replicates"]
        + [f"n_{s}" for s in stage_cols]
        + ["unmapped", "pooled_total", "pooled_staged"]
        + [f"pct_{s}" for s in stage_cols]
        + ["stage_index_mean", "stage_index_sd", "pooled_stage_index",
           "n_animals_mean", "n_animals_sd",
           "survival_pct_mean", "survival_pct_sd", "pooled_survival_pct"]
    )
    cg = _Grid(cond_headers)
    ws = wb.create_sheet("per_condition")
    _write_header(ws, cond_headers)
    for i, row in enumerate(cond_rows):
        r = i + 2
        key = (row["timepoint_h"], row["condition"])
        plo, phi = plate_span[key]
        if row["replicate_unit"] == "plate":
            rep_sheet, rep_grid, (rlo, rhi) = "per_plate", pg, plate_span[key]
        else:
            rep_sheet, rep_grid, (rlo, rhi) = "per_image", ig, cond_img_span[key]
        si_rng = _sheet_ref(rep_sheet, rep_grid.col("stage_index"), rlo, rhi)
        sv_rng = _sheet_ref(rep_sheet, rep_grid.col("survival_pct"), rlo, rhi)
        tot_rng = _sheet_ref(rep_sheet, rep_grid.col("total"), rlo, rhi)
        n_first, n_last = cg.col(f"n_{stage_cols[0]}"), cg.col(f"n_{stage_cols[-1]}")
        pooled_total = f"{cg.col('pooled_total')}{r}"
        pooled_staged = f"{cg.col('pooled_staged')}{r}"
        idx_expr = "+".join(
            f"{w:g}*{cg.col('n_' + s)}{r}" for s, w in indexed) or "0"
        staged_expr = "+".join(
            f"{cg.col('n_' + s)}{r}" for s in indexed_names) or "0"
        surv_expr = "+".join(
            f"{cg.col('n_' + s)}{r}" for s in surv_stages) or "0"
        non_expr = "+".join(
            f"{cg.col('n_' + s)}{r}" for s in non_stages) or "0"

        formulas = {
            "n_plates": f"=COUNT({_sheet_ref('per_plate', pg.col('total'), plo, phi)})",
            "n_images": f"=SUM({_sheet_ref('per_plate', pg.col('n_images'), plo, phi)})",
            "n_replicates": f"=COUNT({tot_rng})",
            "unmapped": f"=SUM({_sheet_ref('per_plate', pg.col('unmapped'), plo, phi)})",
            "pooled_total": f"=SUM({n_first}{r}:{n_last}{r})",
            "pooled_staged": f"={staged_expr}",
            "stage_index_mean": f"=IF(COUNT({si_rng})=0,\"\",AVERAGE({si_rng}))",
            "stage_index_sd": f"=IF(COUNT({si_rng})<2,\"\",STDEV({si_rng}))",
            "pooled_stage_index": (f"=IF({pooled_staged}=0,\"\","
                                   f"({idx_expr})/{pooled_staged})"),
            "n_animals_mean": f"=IF(COUNT({tot_rng})=0,\"\",AVERAGE({tot_rng}))",
            "n_animals_sd": f"=IF(COUNT({tot_rng})<2,\"\",STDEV({tot_rng}))",
            "survival_pct_mean": f"=IF(COUNT({sv_rng})=0,\"\",AVERAGE({sv_rng}))",
            "survival_pct_sd": f"=IF(COUNT({sv_rng})<2,\"\",STDEV({sv_rng}))",
            "pooled_survival_pct": (
                f"=IF({surv_expr}+{non_expr}=0,\"\","
                f"100*({surv_expr})/({surv_expr}+{non_expr}))"),
        }
        for st in stage_cols:
            formulas[f"n_{st}"] = (
                f"=SUM({_sheet_ref('per_plate', pg.col(st), plo, phi)})")
            formulas[f"pct_{st}"] = (
                f"=IF({pooled_total}=0,\"\","
                f"100*{cg.col('n_' + st)}{r}/{pooled_total})")
        out = []
        for h in cond_headers:
            if h in formulas:
                out.append(formulas[h])
            else:
                v = row.get(h)
                out.append("" if v is None else v)
        ws.append(out)
        for h in formulas:
            expected[f"per_condition!{cg.col(h)}{r}"] = row[h]
    _style_body(ws, len(cond_rows), len(cond_headers))
    _autosize(ws, cond_headers, {"condition": 16, "replicate_unit": 15})

    # ---- qc ---------------------------------------------------------------
    qc_rows = agg["qc"]
    qc_headers = [
        "timepoint_h", "condition", "strain", "dose", "unit",
        "n_plates", "n_images", "n_animals_total",
        "animals_per_plate_mean", "animals_per_plate_sd",
        "animals_per_plate_min", "animals_per_plate_max",
        "pct_of_control", "quadrant_cv_pct_mean", "quadrant_cv_pct_max",
        "n_image_errors",
    ]
    qg = _Grid(qc_headers)
    ws = wb.create_sheet("qc")
    _write_header(ws, qc_headers)
    # per_condition row index for each (tp, condition), to point formulas at
    cond_row_of = {(r["timepoint_h"], r["condition"]): i + 2
                   for i, r in enumerate(cond_rows)}
    # the control (lowest dose of the same strain at the same timepoint)
    control_of: dict[tuple, tuple] = {}
    for r in cond_rows:
        if r["dose"] is None:
            continue
        k = (r["timepoint_h"], str(r["strain"]))
        cur = control_of.get(k)
        if cur is None or r["dose"] < cur[0]:
            control_of[k] = (r["dose"], (r["timepoint_h"], r["condition"]))

    qc_row_of = {(q["timepoint_h"], q["condition"]): i + 2
                 for i, q in enumerate(qc_rows)}
    for i, row in enumerate(qc_rows):
        r = i + 2
        key = (row["timepoint_h"], row["condition"])
        plo, phi = plate_span[key]
        cr = cond_row_of[key]
        tot_rng = _sheet_ref("per_plate", pg.col("total"), plo, phi)
        cv_rng = _sheet_ref("per_plate", pg.col("quadrant_cv_pct"), plo, phi)
        ctrl = control_of.get((row["timepoint_h"], str(row["strain"])))
        formulas = {
            "n_plates": f"='per_condition'!{cg.col('n_plates')}{cr}",
            "n_images": f"='per_condition'!{cg.col('n_images')}{cr}",
            "n_animals_total": f"='per_condition'!{cg.col('pooled_total')}{cr}",
            "animals_per_plate_mean": f"=AVERAGE({tot_rng})",
            "animals_per_plate_sd": (f"=IF(COUNT({tot_rng})<2,\"\","
                                     f"STDEV({tot_rng}))"),
            "animals_per_plate_min": f"=MIN({tot_rng})",
            "animals_per_plate_max": f"=MAX({tot_rng})",
            "quadrant_cv_pct_mean": (f"=IF(COUNT({cv_rng})=0,\"\","
                                     f"AVERAGE({cv_rng}))"),
            "quadrant_cv_pct_max": (f"=IF(COUNT({cv_rng})=0,\"\","
                                    f"MAX({cv_rng}))"),
        }
        if ctrl is not None and ctrl[1] in qc_row_of:
            # Point at the control's own qc row, not at per_condition: qc is
            # plate-based and per_condition's n_animals_mean is replicate-based,
            # so mixing them silently divides a plate count by a quadrant count.
            ref = f"'qc'!{qg.col('animals_per_plate_mean')}{qc_row_of[ctrl[1]]}"
            formulas["pct_of_control"] = (
                f"=IF(N({ref})=0,\"\",100*{qg.col('animals_per_plate_mean')}{r}"
                f"/{ref})")
        out = []
        for h in qc_headers:
            if h in formulas:
                out.append(formulas[h])
            else:
                v = row.get(h)
                out.append("" if v is None else v)
        ws.append(out)
        for h in formulas:
            expected[f"qc!{qg.col(h)}{r}"] = row[h]
    _style_body(ws, len(qc_rows), len(qc_headers))
    _autosize(ws, qc_headers, {"condition": 16})

    # ---- gaps (only when there are any) -----------------------------------
    if agg["gaps"]:
        ws = wb.create_sheet("gaps")
        _write_header(ws, ["timepoint_h", "condition", "note"])
        for g in agg["gaps"]:
            ws.append([g["timepoint_h"], g["condition"],
                       "condition present at another timepoint but absent here "
                       "— an empty cell in the figures, not an error"])
        _style_body(ws, len(agg["gaps"]), 3)
        _autosize(ws, ["timepoint_h", "condition", "note"], {"note": 90})

    # ---- size sheets ------------------------------------------------------
    if size:
        _write_size_sheets(wb, size, expected)

    _write_number_formats(wb, stage_cols, ig, pg, cg, qg,
                          len(img_rows), len(plate_rows), len(cond_rows),
                          len(qc_rows))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    write_log(f"Wrote {out_path} ({len(expected)} computed cell(s) to verify)")
    return expected


# index_map is a module-level indirection so write_workbook's signature stays
# readable; survival.py sets it once at import.
index_map: dict = {}


def index_map_lookup(stage: str) -> Optional[float]:
    return index_map.get(str(stage).strip().lower())


def _size_scale_rows(size) -> list:
    """Provenance for the body-size unit. A magnification change mid-experiment
    silently widens every distribution, so the scale that was actually used is
    recorded here rather than inferred from the axis label."""
    if not size:
        return []
    unit = size.get("unit", "px")
    rows = [("size_metric",
             "sqrt(w*h) of the detection box — APPARENT size, not a body "
             "length: a coiled animal reads smaller than a straight one of the "
             "same length"),
            ("size_unit", "µm" if unit == "um" else "px"),
            ("size_unit_reason", size.get("scale_note", ""))]
    if unit == "um":
        lo, mid, hi = (size.get("um_per_px_min"), size.get("um_per_px_median"),
                       size.get("um_per_px_max"))
        if lo is not None and hi is not None:
            spread = 100.0 * (hi / lo - 1.0) if lo else 0.0
            rows.append(("um_per_px",
                         f"{lo:.4g}–{hi:.4g} (median {mid:.4g}), read from each "
                         f"image's TIFF tags; spread {spread:.1f}%"
                         + ("" if spread < 1.0 else
                            " — the working distance moved during this "
                            "experiment, so sizes are comparable only to the "
                            "extent that tag is right")))
    return rows


def _run_info_rows(meta, plans, summary, stage_names, unmapped, agg, size=None):
    # Local import: survival.py imports this module lazily, inside the run.
    from survival import (SURVIVAL_CONFIG, _RESULTS_NAME, _MODEL_PATH,
                          _VISION_PY)
    eff_conf = meta.get("class_conf") or {}
    seam = meta.get("seam") or {}
    rescore = meta.get("rescore") or {}
    alpha = float(rescore.get("alpha") or 0.0)
    rows = [
        ("timestamp", datetime.now().isoformat(timespec="seconds")),
        ("workbook", _RESULTS_NAME),
        ("model_path", str(_MODEL_PATH)),
        ("conf_min", meta.get("conf", "")),
        ("conf_per_class",
         ", ".join(f"{k}={float(v):.2f}" for k, v in eff_conf.items())
         or "(uniform)"),
        ("tile", f"676x608 (tiled_infer, overlap {meta.get('overlap', '?')})"),
        ("seam_suppression",
         f"margin {seam.get('margin_px')} px, cover {seam.get('cover_frac')}"
         if seam.get("cover_frac") is not None else "off"),
        ("class_agnostic_iou", meta.get("class_agnostic_iou") or "off"),
        ("class_size_px",
         ", ".join(f"{k}={list(v)}" for k, v in
                   (meta.get("class_size_px") or {}).items()) or "off"),
        ("rescore",
         (f"alpha {alpha:g} — class-confidence correction ON (relabels only, "
          "detection count unchanged); refs "
          + ", ".join(f"{k}={float(v):.4f}"
                      for k, v in (rescore.get("refs") or {}).items()))
         if alpha else "off (alpha 0 — arg-max on raw scores)"),
        ("excluded_classes",
         ", ".join(meta.get("exclude_classes") or []) or "(none)"),
        ("excluded_note",
         "excluded classes were NOT detected — their absence is not a count of 0"
         if meta.get("exclude_classes") else ""),
        ("n_folders", len(plans)),
    ]
    for p in plans:
        rows.append((f"folder[{p.hours:g} h]", f"{p.folder} — {p.detail}"))
    rows += [
        ("timepoint_h_values",
         ", ".join(f"{t:g}" for t in agg["timepoints"])),
        ("grouping_mode", summary.get("mode", "")),
        ("encoded_fraction", f"{summary.get('encoded_fraction', 0.0):.2f}"),
        ("n_conditions", summary.get("n_conditions", 0)),
        ("n_plates", summary.get("n_plates", 0)),
        ("image_count", summary.get("n_images", 0)),
        ("n_image_errors", agg["n_error"]),
        ("n_unparsed_images", agg["n_unparsed"]),
        ("images_reused",
         f"{summary.get('n_reused', 0)} (detections read back from earlier "
         "runs, not re-analysed)"),
        ("images_analysed_now", summary.get("n_analysed", 0)),
        ("cached_detections_relabelled",
         f"{summary.get('n_relabelled', 0)} (re-argmaxed at the current alpha "
         "from the saved per-class scores — exact, not an approximation)"
         if summary.get("n_relabelled") else 0),
        ("model_run_this_time",
         "no — every detection came from a previous run"
         if summary.get("from_cache_only") else "yes"),
        ("n_condition_gaps", len(agg["gaps"])),
        ("model_classes", ", ".join(stage_names)),
        ("stage_index_mapping",
         "L1=1, L2=2, L3=3, L4=4, adult=5; eggs excluded"),
        ("survivors", ", ".join(SURVIVAL_CONFIG["survivors"])),
        ("non_survivors", ", ".join(SURVIVAL_CONFIG["non_survivors"])),
        ("excluded", ", ".join(SURVIVAL_CONFIG["excluded"])),
        ("unmapped_stages", ", ".join(unmapped) if unmapped else "(none)"),
        ("survival_formula",
         "survivors / (survivors + non_survivors) * 100 — workbook only, in no "
         "figure; see the README sheet"),
        ("vision_python", str(_VISION_PY)),
    ]
    rows += _size_scale_rows(size)
    return rows


def _write_size_sheets(wb, size: dict, expected: dict) -> None:
    """size_histogram (the exact bins the body-size figure draws) + percentiles.

    Values, not formulas: they are derived from soft_stage_scores.csv, whose
    per-animal rows are deliberately not in this workbook. The one live cell is
    size_summary's n, which sums its own histogram column — so the two sheets
    disagreeing is visible rather than silent.
    """
    from openpyxl.styles import Font

    keys = list(size["groups"].keys())
    edges = size["bin_edges"]
    u = size.get("unit", "px")          # "um" or "px" — set by survival_size
    ws = wb.create_sheet("size_histogram")
    headers = [f"bin_lo_{u}", f"bin_hi_{u}", f"bin_mid_{u}"] + keys
    _write_header(ws, headers)
    for i in range(len(edges) - 1):
        row = [round(edges[i], 2), round(edges[i + 1], 2),
               round((edges[i] * edges[i + 1]) ** 0.5, 2)]
        for k in keys:
            row.append(size["groups"][k]["hist"][i])
        ws.append(row)
    n_bins = len(edges) - 1
    total_row = n_bins + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(
        name=_FONT, size=10, bold=True)
    for j, k in enumerate(keys):
        c = _col(4 + j)
        ws.cell(row=total_row, column=4 + j,
                value=f"=SUM({c}2:{c}{n_bins + 1})").font = Font(
            name=_FONT, size=10, bold=True)
        expected[f"size_histogram!{c}{total_row}"] = size["groups"][k]["n"]
    _style_body(ws, n_bins, len(headers))
    _autosize(ws, headers)
    ws.column_dimensions["A"].width = 12

    ws = wb.create_sheet("size_summary")
    # "unit" here is the DOSE unit (J/m², µM) and has always been; "size_unit"
    # is the new one and says what the percentile columns are measured in.
    sh = ["group", "strain", "dose", "unit", "timepoint_h", "n", "size_unit",
          "p10", "p25", "p50_median", "p75", "p90", "mean", "geometric_mean"]
    _write_header(ws, sh)
    size_unit = "µm" if u == "um" else "px"
    for j, k in enumerate(keys):
        g = size["groups"][k]
        hist_col = _col(4 + j)
        r = j + 2
        ws.append([
            k, g["strain"], g["dose"], g["unit"], g["timepoint_h"],
            f"=SUM('size_histogram'!{hist_col}$2:{hist_col}${n_bins + 1})",
            size_unit,
            g["p10"], g["p25"], g["p50"], g["p75"], g["p90"],
            g["mean"], g["gmean"],
        ])
        expected[f"size_summary!F{r}"] = g["n"]
    _style_body(ws, len(keys), len(sh))
    _autosize(ws, sh, {"group": 26})


def _write_number_formats(wb, stage_cols, ig, pg, cg, qg,
                          n_img, n_plate, n_cond, n_qc) -> None:
    def fmt(sheet, grid, headers, n_rows, code):
        ws = wb[sheet]
        for h in headers:
            if h not in grid.pos:
                continue
            c = grid.col(h)
            for r in range(2, n_rows + 2):
                ws[f"{c}{r}"].number_format = code

    fmt("per_image", ig, ["stage_index"], n_img, _F_2)
    fmt("per_image", ig, ["survival_pct", "timepoint_h"], n_img, _F_1)
    fmt("per_plate", pg, ["stage_index"], n_plate, _F_2)
    fmt("per_plate", pg, ["survival_pct", "timepoint_h", "quadrant_cv_pct"],
        n_plate, _F_1)
    fmt("per_condition", cg,
        ["stage_index_mean", "stage_index_sd", "pooled_stage_index"],
        n_cond, _F_2)
    fmt("per_condition", cg,
        [f"pct_{s}" for s in stage_cols]
        + ["survival_pct_mean", "survival_pct_sd", "pooled_survival_pct",
           "n_animals_mean", "n_animals_sd", "timepoint_h"], n_cond, _F_1)
    fmt("qc", qg,
        ["animals_per_plate_mean", "animals_per_plate_sd", "pct_of_control",
         "quadrant_cv_pct_mean", "quadrant_cv_pct_max", "timepoint_h"],
        n_qc, _F_1)


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

def verify_workbook(path: Path, expected: dict,
                    write_log: Callable[[str], None],
                    tol: float = 5e-6) -> list[str]:
    """Re-read the RECALCULATED workbook and compare every computed cell.

    A clean recalculation proves the formulas evaluate. It does not prove they
    are right: an off-by-one row range recalculates without error and reports
    the wrong number. This is the check that catches that. Returns a list of
    human-readable mismatches (empty = clean).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    problems: list[str] = []
    n_checked = 0
    for addr, want in expected.items():
        sheet, cell = addr.split("!", 1)
        if sheet not in wb.sheetnames:
            problems.append(f"{addr}: sheet missing")
            continue
        got = wb[sheet][cell].value
        n_checked += 1
        want_nan = isinstance(want, float) and want != want
        if want_nan or want is None:
            # Python has no value here; the cell must be blank, not a number.
            if got not in (None, ""):
                problems.append(f"{addr}: expected blank, got {got!r}")
            continue
        if got is None or isinstance(got, str):
            problems.append(f"{addr}: expected {want!r}, got {got!r}")
            continue
        if abs(float(got) - float(want)) > tol * max(1.0, abs(float(want))):
            problems.append(f"{addr}: expected {want!r}, got {got!r}")
    wb.close()
    if problems:
        write_log(f"WORKBOOK CROSS-CHECK FAILED: {len(problems)} of "
                  f"{n_checked} computed cell(s) disagree with Python:")
        for p in problems[:40]:
            write_log(f"    {p}")
        if len(problems) > 40:
            write_log(f"    … and {len(problems) - 40} more")
    else:
        write_log(f"Workbook cross-check: {n_checked} computed cell(s) match "
                  "the Python values exactly.")
    return problems
